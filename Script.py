from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium import webdriver
import time
import openpyxl

driver = webdriver.Chrome()
driver.get('https://portalbi.trf5.jus.br/portal-bi/painel.html?id=3002')
driver.maximize_window()


caminho_planilha = 'ProcessosMigrados.xlsx'
planilha = load_workbook(caminho_planilha)
planilha_ativa = planilha.active


workbook = openpyxl.Workbook()
sheet_processos = planilha_ativa.create_sheet('Informações')
sheet_processos['A1'].value = 'Sistema'
sheet_processos['B1'].value = 'Seção'


for linha in range(1, planilha_ativa.max_row + 1):

    informacao_excel = planilha_ativa.cell(row=linha, column=1).value

    pesquisar_processos = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="jUVcvU_content"]/div/div/input'))
    )
    pesquisar_processos.clear()
    pesquisar_processos.send_keys(informacao_excel)

    acessar_processos = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="UcJuc_content"]/div/button/text/span'))
    )
    acessar_processos.click()

    sistema = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, '//*[@id="PmHLLh_content"]/div/div[2]/div[1]/div/table/tbody/tr[1]/td[2]/div/div/span'))
).text

    secao = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, '//*[@id="PmHLLh_content"]/div/div[2]/div[1]/div/table/tbody/tr[1]/td[3]/div/div/span'))
).text

    sheet_processos.append([sistema, secao])


    for sistema, secao in zip(sistema, secao):
        sheet_processos.append([sistema.text, secao.text])

    salvar_fechar = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="lawsuit-edit-form"]/div[24]/div/div/div/button'))
    )
    salvar_fechar.click()

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="menuProcessos"]'))
    )


workbook.save('ProcessosTribunal.xlsx')


driver.quit()
